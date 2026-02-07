#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
mapping_tool.py
==============
Outil unique pour gérer le mapping GEDCOM -> TMG.

Usage:
  python mapping_tool.py G   ->  Génère l'Excel avec suggestions intelligentes
  python mapping_tool.py C   ->  Lit l'Excel validé -> produit mapping.json

Architecture:
  Excel = interface unique (vous ne touchez que ça)
  JSON  = output technique (généré automatiquement, jamais édité à la main)

Hiérarchie des suggestions (ordre de priorité):
  1. mapping.json précédent (dernière validation utilisateur)
  2. HOST de D.DBF (mapping TMG natif — ce que l'utilisateur a fait à l'import)
  3. Match exact après normalize() dans T.DBF
  4. Table de standards GEDCOM universels
  (si rien ne matche -> rouge, à remplir manuellement)

Note: HOST ne contient QUE les events. Les rôles ne bénéficient pas de cette source.
Portabilité: tout utilisateur TMG qui a importé un GEDCOM a son propre HOST.

Règle fondamentale:
  Clé = normalize(tag) = NFKD + ASCII + UPPER + STRIP
  UNE seule fonction, utilisée partout. Pas de double circuit.
"""

import sys
import os
import json
import re
import unicodedata
from collections import defaultdict

try:
    import dbf
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install dbf.py openpyxl")
    sys.exit(1)

# =============================================================================
# CONFIGURATION — GUI uniquement, rien de hardcodé
# =============================================================================
import tkinter as tk
from tkinter import filedialog

GEDCOM_PATH  = None
TMG_PATH     = None
TMG_PREFIX   = None
EXCEL_FILE   = "mapping_master.xlsx"
JSON_FILE    = "mapping.json"
LOG_CALLBACK = None  # Callback optionnel pour logs vers GUI

def log(message, level='INFO'):
    """
    Log message - utilise callback si fourni, sinon print()
    
    Usage CLI: log("message") → print()
    Usage GUI: log("message") → callback(message, level)
    """
    if LOG_CALLBACK:
        LOG_CALLBACK(message, level)
    else:
        print(message)

def _extract_prefix(pjc_path):
    """
    Déduit le préfixe depuis le nom du fichier .PJC.
    Convention TMG: testcase__.PJC  ->  préfixe = testcase_
    Règle: enlever .PJC puis le dernier underscore.
    """
    basename = os.path.splitext(os.path.basename(pjc_path))[0]  # testcase__
    if basename.endswith('_'):
        prefix = basename[:-1]  # testcase_
    else:
        prefix = basename + '_'
    # Vérification: D.DBF avec ce préfixe existe?
    tmg_dir = os.path.dirname(pjc_path)
    ddbf = os.path.join(tmg_dir, f"{prefix}D.DBF")
    if not os.path.exists(ddbf):
        return None
    return prefix

def select_paths_gui():
    """Ouvre les dialogues GUI pour choisir GEDCOM + projet TMG."""
    global GEDCOM_PATH, TMG_PATH, TMG_PREFIX

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    # 1. Sélection GEDCOM
    log("\n🖱️  Select your GEDCOM file...")
    GEDCOM_PATH = filedialog.askopenfilename(
        title="Select GEDCOM file",
        filetypes=[("GEDCOM", "*.ged"), ("All", "*.*")]
    )
    if not GEDCOM_PATH:
        log("❌ Cancelled.", 'ERROR')
        sys.exit(0)
    log(f"   ✓ GEDCOM: {GEDCOM_PATH}", 'SUCCESS')

    # 2. Sélection projet TMG (.PJC)
    log("🖱️  Select your TMG project file (.PJC)...")
    pjc_path = filedialog.askopenfilename(
        title="Select TMG project file (.PJC)",
        filetypes=[("TMG Project", "*.PJC"), ("All", "*.*")]
    )
    if not pjc_path:
        log("❌ Cancelled.", 'ERROR')
        sys.exit(0)

    TMG_PATH = os.path.dirname(pjc_path)
    log(f"   ✓ Project: {pjc_path}", 'SUCCESS')

    root.destroy()

    # 3. Détection préfixe depuis le nom du .PJC
    TMG_PREFIX = _extract_prefix(pjc_path)
    if not TMG_PREFIX:
        log("❌ Cannot determine project prefix.", 'ERROR')
        sys.exit(1)
    log(f"   ✓ Prefix: {TMG_PREFIX}", 'SUCCESS')

    # Vérification D.DBF
    ddbf = os.path.join(TMG_PATH, f"{TMG_PREFIX}D.DBF")
    if not os.path.exists(ddbf):
        log(f"❌ {ddbf} not found.", 'ERROR')
        sys.exit(1)

# =============================================================================
# NORMALISATION — UNE SEULE FONCTION, UTILISÉE PARTOUT
# =============================================================================
def normalize(txt):
    if not txt:
        return ""
    nfkd = unicodedata.normalize('NFKD', str(txt))
    ascii_str = nfkd.encode('ascii', 'ignore').decode('ascii')
    return ascii_str.upper().strip()

def get_tmg_file(suffix):
    return os.path.join(TMG_PATH, f"{TMG_PREFIX}{suffix}.DBF")

# =============================================================================
# STYLES EXCEL
# =============================================================================
FILL_HDR    = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
FONT_HDR    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
FILL_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GREY   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FONT_NORMAL = Font(name="Arial", size=10)
FONT_BOLD   = Font(name="Arial", size=10, bold=True)
ALIGN_C     = Alignment(horizontal="center", vertical="center")
BORDER      = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)

def style_header_row(ws, n_cols):
    for col in range(1, n_cols + 1):
        c = ws.cell(1, col)
        c.font = FONT_HDR
        c.fill = FILL_HDR
        c.alignment = ALIGN_C
        c.border = BORDER

def style_data_row(ws, row_idx, n_cols, fill):
    for col in range(1, n_cols + 1):
        c = ws.cell(row_idx, col)
        c.font = FONT_NORMAL
        c.fill = fill
        c.border = BORDER

# =============================================================================
# 1. MÉMOIRE — extraction de HOST (D.DBF) et/ou mapping.json précédent
# =============================================================================
def load_host_ddbf():
    """
    Lit le champ memo HOST de D.DBF — table VFP binaire embarquée.
    Retourne un dict: {normalize(NAME): NAMENEW_exact, ...}
    
    Protocole de lecture:
      1. D.DBF: champ HOST (type M, 4 bytes LE) = numéro de bloc dans D.FPT
      2. D.FPT: block_size = file_size / next_block (offset 0, 4 bytes BE)
      3. Memo header: type(4 BE) + length(4 BE)
      4. Memo data = table VFP standard (version 0x30)
      5. Parser champs NAME et NAMENEW depuis les descripteurs
      6. Lecture séquentielle des records (champs C, fixed-width, latin-1)
    """
    import struct

    d_dbf = os.path.join(TMG_PATH, f"{TMG_PREFIX}D.DBF")
    d_fpt = os.path.join(TMG_PATH, f"{TMG_PREFIX}D.FPT")
    host_map = {}

    if not os.path.exists(d_dbf) or not os.path.exists(d_fpt):
        log("   ⚠️  D.DBF/D.FPT not found — HOST unavailable", 'WARNING')
        return host_map

    try:
        with open(d_dbf, 'rb') as f:
            dbf_data = f.read()
        with open(d_fpt, 'rb') as f:
            fpt_data = f.read()

        # --- FPT: block_size ---
        next_block = struct.unpack_from('>I', fpt_data, 0)[0]
        if next_block == 0:
            log("   ⚠️  Empty FPT (next_block=0)", 'WARNING')
            return host_map
        block_size = len(fpt_data) // next_block

        # --- D.DBF: trouver le champ HOST ---
        dbf_hdr_size = struct.unpack_from('<H', dbf_data, 8)[0]

        # Parse descripteurs pour trouver l'offset de HOST dans le record
        fields_dbf = []
        off = 32
        while off < dbf_hdr_size and dbf_data[off] != 0x0D:
            fname = dbf_data[off:off+11].split(b'\x00')[0].decode('ascii', errors='replace')
            ftype = chr(dbf_data[off+11])
            flen  = dbf_data[off+16]
            fields_dbf.append({'name': fname, 'type': ftype, 'len': flen})
            off += 32

        pos = 1
        for f in fields_dbf:
            f['offset'] = pos
            pos += f['len']

        host_field = next((f for f in fields_dbf if f['name'] == 'HOST'), None)
        if not host_field:
            log("   ⚠️  HOST field not found in D.DBF", 'WARNING')
            return host_map

        # Bloc memo HOST (int32 LE dans le premier record)
        host_block = struct.unpack_from('<I', dbf_data, dbf_hdr_size + host_field['offset'])[0]
        if host_block == 0:
            log("   ⚠️  Empty HOST memo (block=0)", 'WARNING')
            return host_map

        # --- FPT: lire le memo ---
        memo_offset = host_block * block_size
        memo_type   = struct.unpack_from('>I', fpt_data, memo_offset)[0]
        memo_len    = struct.unpack_from('>I', fpt_data, memo_offset + 4)[0]

        if memo_type != 1:
            log(f"   ⚠️  Unexpected memo type: {memo_type}", 'WARNING')
            return host_map

        memo = fpt_data[memo_offset + 8 : memo_offset + 8 + memo_len]

        # --- Parse table VFP embarquée ---
        num_recs   = struct.unpack_from('<I', memo, 4)[0]
        tbl_hdr_sz = struct.unpack_from('<H', memo, 8)[0]
        rec_size   = struct.unpack_from('<H', memo, 10)[0]

        # Descripteurs de champs de la table embarquée
        fields_tbl = []
        off = 32
        while off < tbl_hdr_sz and memo[off] != 0x0D:
            fname = memo[off:off+11].split(b'\x00')[0].decode('ascii', errors='replace')
            ftype = chr(memo[off+11])
            flen  = memo[off+16]
            fields_tbl.append({'name': fname, 'type': ftype, 'len': flen})
            off += 32

        pos = 1
        for f in fields_tbl:
            f['offset'] = pos
            pos += f['len']

        name_f    = next((f for f in fields_tbl if f['name'] == 'NAME'), None)
        namenew_f = next((f for f in fields_tbl if f['name'] == 'NAMENEW'), None)

        if not name_f or not namenew_f:
            log("   ⚠️  NAME/NAMENEW fields not found in HOST table", 'WARNING')
            return host_map

        # --- Lecture des records ---
        for i in range(num_recs):
            base = tbl_hdr_sz + i * rec_size
            name    = memo[base + name_f['offset']    : base + name_f['offset']    + name_f['len']].decode('latin-1').strip()
            namenew = memo[base + namenew_f['offset'] : base + namenew_f['offset'] + namenew_f['len']].decode('latin-1').strip()
            if name and namenew:
                host_map[normalize(name)] = namenew

        log(f"   🔓 HOST lu depuis D.DBF: {len(host_map)} associations", 'SUCCESS')
        return host_map

    except Exception as e:
        log(f"   ⚠️  HOST read error: {e}", 'WARNING')
        return host_map


def load_memory_json():
    """Charge mapping.json précédent s'il existe."""
    events_json = {}
    roles_json  = {}
    if not os.path.exists(JSON_FILE):
        return events_json, roles_json

    log(f"   💾 Reading previous mapping.json...", 'INFO')
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    for k, v in data.get('events', {}).items():
        if v.get('tmg_name'):
            events_json[normalize(k)] = v['tmg_name']
    for k, v in data.get('roles', {}).items():
        if v.get('eng'):
            roles_json[normalize(k)] = {'eng': v['eng'], 'fra': v.get('fra', v['eng'])}
    return events_json, roles_json


# =============================================================================
# 2. SCAN GEDCOM — events avec _SHAR + rôles
# =============================================================================
def scan_gedcom():
    """
    Retourne:
      events: {normalize(key): {'raw': str, 'freq': int}}
      roles:  {normalize(key): {'raw': str, 'freq': int}}
    """
    events = {}
    roles  = {}

    # Première passe: parser les blocs pour capturer events avec _SHAR
    current_evt_key = None  # clé normalisée de l'event en cours
    current_evt_raw = None  # forme originale
    evt_has_shar = False

    with open(GEDCOM_PATH, 'r', encoding='ansi') as f:
        for line in f:
            p = line.strip().split(' ', 2)
            if len(p) < 2:
                continue
            lvl = p[0]
            tag = p[1]
            val = p[2].strip() if len(p) > 2 else ""

            if lvl == '0':
                current_evt_key = None
                current_evt_raw = None
                evt_has_shar = False

            elif lvl == '1':
                current_evt_key = None
                current_evt_raw = None
                evt_has_shar = False

                if tag in ('EVEN', 'FACT'):
                    current_evt_key = '__PENDING__'
                    current_evt_raw = tag
                else:
                    current_evt_key = normalize(tag)
                    current_evt_raw = tag

            elif lvl == '2':
                if tag == 'TYPE' and current_evt_key == '__PENDING__':
                    current_evt_key = normalize(val)
                    current_evt_raw = val

                elif tag == '_SHAR' and current_evt_key and current_evt_key != '__PENDING__':
                    evt_has_shar = True
                    nk = current_evt_key
                    if nk not in events:
                        events[nk] = {'raw': current_evt_raw, 'freq': 0}
                    events[nk]['freq'] += 1

            elif lvl == '3':
                if tag == 'ROLE' and evt_has_shar and current_evt_key:
                    nk = normalize(val)
                    if nk not in roles:
                        roles[nk] = {'raw': val, 'freq': 0}
                    roles[nk]['freq'] += 1

    return events, roles


# =============================================================================
# 3. SCAN T.DBF — events + rôles avec labels EN/FR
# =============================================================================
def scan_tdbf():
    """
    Retourne:
      events_tmg: {normalize(name): {'name': str, 'etype': int}}
      roles_tmg:  {normalize(eng_label): {'eng': str, 'fra': str, 'code': str}}
    """
    events_tmg = {}
    roles_tmg  = {}

    regex_role  = re.compile(r'\[RL=(\d+)\](.*?)(?=\[RL=|\[:LABELS\]|$)', re.DOTALL)
    regex_label = re.compile(r'\[L=([^\]]+)\]([^\[\r\n]+)')

    t = dbf.Table(get_tmg_file("T"))
    t.open()
    for rec in t:
        name  = str(rec.ETYPENAME).strip()
        etype = rec.ETYPENUM
        sent  = str(rec.TSENTENCE)

        events_tmg[normalize(name)] = {'name': name, 'etype': etype}

        if '[LABELS:]' not in sent:
            continue
        try:
            labels_section = sent.split('[LABELS:]')[1].split('[:LABELS]')[0]
            for m in regex_role.finditer(labels_section):
                code  = m.group(1)
                block = m.group(2)
                eng = fra = None
                for lm in regex_label.finditer(block):
                    lang  = lm.group(1).upper()
                    label = lm.group(2).strip()
                    if lang == 'ENGLISH':
                        eng = label
                    elif lang == 'FRENCH':
                        fra = label
                if eng:
                    norm_eng = normalize(eng)
                    if norm_eng not in roles_tmg:
                        roles_tmg[norm_eng] = {'eng': eng, 'fra': fra or eng, 'code': code}
        except Exception:
            pass
    t.close()
    return events_tmg, roles_tmg


# =============================================================================
# 4. LOGIQUE DE SUGGESTION
# =============================================================================

# Standards GEDCOM universels (utilisés seulement si pas de mémoire)
GEDCOM_STANDARDS = {
    "BIRT": "Birth",
    "DEAT": "Death",
    "MARR": "Marriage",
    "BURI": "Burial",
    "BAPM": "Baptism",
    "MARC": "Marriage contract",
    "OCCU": "Occupation",
    "RESI": "Residence",
}


def suggest_event(ged_key, events_host, events_json, events_tmg):
    """
    Retourne: (suggestion, origine, note)
    suggestion = nom TMG exact à utiliser (ou None)
    origine = "JSON prev" / "HOST" / "Match TMG" / "Standard" / None
    note = explication si problème
    """
    # Priorité 1: JSON précédent (dernière validation utilisateur)
    if ged_key in events_json:
        tmg_name = events_json[ged_key]
        if normalize(tmg_name) in events_tmg:
            return events_tmg[normalize(tmg_name)]['name'], "JSON prev", ""

    # Priorité 2: HOST de D.DBF (mapping TMG natif)
    if ged_key in events_host:
        host_target = events_host[ged_key]
        if normalize(host_target) in events_tmg:
            return events_tmg[normalize(host_target)]['name'], "HOST", ""
        else:
            return None, None, f"HOST disait '{host_target}' — inconnu dans TMG"

    # Priorité 3: Match exact TMG
    if ged_key in events_tmg:
        return events_tmg[ged_key]['name'], "Match TMG", ""

    # Priorité 4: Standards GEDCOM
    if ged_key in GEDCOM_STANDARDS:
        std_target = GEDCOM_STANDARDS[ged_key]
        if normalize(std_target) in events_tmg:
            return events_tmg[normalize(std_target)]['name'], "Standard", ""

    return None, None, ""


def suggest_role(ged_key, roles_json, roles_tmg):
    """
    Retourne: (eng, fra, origine, note)
    Note: HOST ne contient pas de rôles. Priorité: JSON prev → Match TMG → Alias FR
    """
    # Priorité 1: JSON précédent
    if ged_key in roles_json:
        info = roles_json[ged_key]
        return info['eng'], info['fra'], "JSON prev", ""

    # Priorité 2: Match direct anglais TMG
    if ged_key in roles_tmg:
        info = roles_tmg[ged_key]
        return info['eng'], info['fra'], "Match TMG", ""

    # Priorité 2b: Match sur français TMG (alias)
    for norm_eng, info in roles_tmg.items():
        if normalize(info['fra']) == ged_key:
            return info['eng'], info['fra'], "Alias FR", ""

    return None, None, None, ""


# =============================================================================
# 5. GÉNÉRATION EXCEL (mode G)
# =============================================================================
def generate_excel(events_ged, roles_ged, events_tmg, roles_tmg,
                   events_host, events_json, roles_json):

    wb = openpyxl.Workbook()

    # =========================================================================
    # ONGLET 1: EVENTS
    #   A: Clé normalisée
    #   B: Tag/Type GED raw
    #   C: Fréquence
    #   D: HOST D.DBF (mapping natif TMG — lecture seule, info)
    #   E: Suggestion auto (meilleur guess — lecture seule)
    #   F: Origine suggestion
    #   G: ★ À VALIDER ★ (pré-rempli avec suggestion. SEULE colonne à éditer)
    #   H: ID TMG (info)
    #   I: Note
    # =========================================================================
    ws = wb.active
    ws.title = "EVENTS"

    headers = [
        "Clé normalisée",
        "Tag GED raw",
        "Freq",
        "HOST (D.DBF)",
        "Suggestion auto",
        "Origine",
        "★ À VALIDER ★",
        "ID TMG",
        "Note"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    style_header_row(ws, len(headers))

    widths = [28, 24, 8, 24, 24, 14, 28, 10, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    stats = {'green': 0, 'yellow': 0, 'red': 0}

    for ged_key in sorted(events_ged.keys()):
        info = events_ged[ged_key]
        host_val = events_host.get(ged_key, "")
        suggestion, origine, note = suggest_event(
            ged_key, events_host, events_json, events_tmg
        )

        # Colonne G = suggestion si dispo, sinon vide (à remplir)
        valider = suggestion or ""

        # ID TMG correspondant
        tmg_id = ""
        if suggestion and normalize(suggestion) in events_tmg:
            tmg_id = events_tmg[normalize(suggestion)]['etype']

        # Couleur selon statut
        if suggestion:
            if origine == "JSON prev":
                fill = FILL_GREEN
                stats['green'] += 1
            else:
                fill = FILL_YELLOW
                stats['yellow'] += 1
        else:
            fill = FILL_RED
            stats['red'] += 1

        ws.cell(row, 1, ged_key)
        ws.cell(row, 2, info['raw'])
        ws.cell(row, 3, info['freq'])
        ws.cell(row, 4, host_val if host_val else "—")
        ws.cell(row, 5, suggestion if suggestion else "—")
        ws.cell(row, 6, origine if origine else "—")
        ws.cell(row, 7, valider)
        ws.cell(row, 8, tmg_id if tmg_id else "—")
        ws.cell(row, 9, note)

        style_data_row(ws, row, len(headers), fill)
        ws.cell(row, 7).font = FONT_BOLD  # ★ colonne à valider en gras

        row += 1

    # Ligne résumé
    ws.cell(row, 1, "RÉSUMÉ").font = FONT_BOLD
    ws.cell(row, 7, f"🟢 {stats['green']} validés | 🟡 {stats['yellow']} à confirmer | 🔴 {stats['red']} à remplir").font = FONT_BOLD
    style_data_row(ws, row, len(headers), FILL_GREY)

    # =========================================================================
    # ONGLET 2: ROLES
    #   A: Clé normalisée
    #   B: Rôle GED raw
    #   C: Fréquence
    #   D: (réservé — HOST events seulement)
    #   E: Suggestion eng
    #   F: Suggestion fra
    #   G: Origine
    #   H: ★ À VALIDER eng ★
    #   I: ★ À VALIDER fra ★
    #   J: Note
    # =========================================================================
    ws2 = wb.create_sheet("ROLES")

    headers2 = [
        "Clé normalisée",
        "Rôle GED raw",
        "Freq",
        "—",
        "Suggestion eng",
        "Suggestion fra",
        "Origine",
        "★ À VALIDER eng ★",
        "★ À VALIDER fra ★",
        "Note"
    ]
    for col, h in enumerate(headers2, 1):
        ws2.cell(1, col, h)
    style_header_row(ws2, len(headers2))

    widths2 = [22, 28, 8, 28, 18, 18, 14, 22, 22, 50]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    row = 2
    stats2 = {'green': 0, 'yellow': 0, 'red': 0}

    for ged_key in sorted(roles_ged.keys()):
        info = roles_ged[ged_key]

        eng, fra, origine, note = suggest_role(
            ged_key, roles_json, roles_tmg
        )

        valider_eng = eng or ""
        valider_fra = fra or ""

        if eng:
            if origine == "JSON prev":
                fill = FILL_GREEN
                stats2['green'] += 1
            else:
                fill = FILL_YELLOW
                stats2['yellow'] += 1
        else:
            fill = FILL_RED
            stats2['red'] += 1

        ws2.cell(row, 1,  ged_key)
        ws2.cell(row, 2,  info['raw'])
        ws2.cell(row, 3,  info['freq'])
        ws2.cell(row, 4,  "—")
        ws2.cell(row, 5,  eng if eng else "—")
        ws2.cell(row, 6,  fra if fra else "—")
        ws2.cell(row, 7,  origine if origine else "—")
        ws2.cell(row, 8,  valider_eng)
        ws2.cell(row, 9,  valider_fra)
        ws2.cell(row, 10, note)

        style_data_row(ws2, row, len(headers2), fill)
        ws2.cell(row, 8).font = FONT_BOLD
        ws2.cell(row, 9).font = FONT_BOLD

        row += 1

    # Résumé rôles
    ws2.cell(row, 1, "RÉSUMÉ").font = FONT_BOLD
    ws2.cell(row, 8, f"🟢 {stats2['green']} | 🟡 {stats2['yellow']} | 🔴 {stats2['red']}").font = FONT_BOLD
    style_data_row(ws2, row, len(headers2), FILL_GREY)

    # =========================================================================
    # ONGLET 3: RÉFÉRENCE TMG
    # Liste complète des events et rôles disponibles dans T.DBF
    # Pour consulter sans sortir du fichier
    # =========================================================================
    ws3 = wb.create_sheet("Référence TMG")

    # Section Events
    ws3.cell(1, 1, "EVENTS DISPONIBLES DANS TMG").font = Font(name="Arial", size=11, bold=True, color="366092")
    headers3a = ["Nom exact", "ID (etype)", "Clé normalisée"]
    for col, h in enumerate(headers3a, 1):
        c = ws3.cell(2, col, h)
        c.font = FONT_HDR
        c.fill = FILL_HDR
        c.alignment = ALIGN_C
        c.border = BORDER

    row = 3
    for norm_key in sorted(events_tmg.keys()):
        info = events_tmg[norm_key]
        ws3.cell(row, 1, info['name'])
        ws3.cell(row, 2, info['etype'])
        ws3.cell(row, 3, norm_key)
        for col in range(1, 4):
            ws3.cell(row, col).font = FONT_NORMAL
            ws3.cell(row, col).border = BORDER
        row += 1

    # Section Rôles (après un gap)
    row += 1
    ws3.cell(row, 1, "RÔLES DISPONIBLES DANS TMG").font = Font(name="Arial", size=11, bold=True, color="366092")
    row += 1
    headers3b = ["Label EN", "Label FR", "Code RL", "Clé normalisée"]
    for col, h in enumerate(headers3b, 1):
        c = ws3.cell(row, col, h)
        c.font = FONT_HDR
        c.fill = FILL_HDR
        c.alignment = ALIGN_C
        c.border = BORDER
    row += 1

    for norm_key in sorted(roles_tmg.keys()):
        info = roles_tmg[norm_key]
        ws3.cell(row, 1, info['eng'])
        ws3.cell(row, 2, info['fra'])
        ws3.cell(row, 3, info['code'])
        ws3.cell(row, 4, norm_key)
        for col in range(1, 5):
            ws3.cell(row, col).font = FONT_NORMAL
            ws3.cell(row, col).border = BORDER
        row += 1

    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 24
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 28

    # =========================================================================
    # ONGLET 4: CONFLITS
    # Plusieurs clés GED vers même event TMG
    # =========================================================================
    ws4 = wb.create_sheet("Conflits")

    ws4.cell(1, 1, "Plusieurs clés GED pointent vers le même event TMG").font = Font(
        name="Arial", size=11, bold=True, color="366092"
    )
    headers4 = ["Event TMG cible", "Clés GED concernées", "Évaluation"]
    for col, h in enumerate(headers4, 1):
        c = ws4.cell(2, col, h)
        c.font = FONT_HDR
        c.fill = FILL_HDR
        c.alignment = ALIGN_C
        c.border = BORDER

    # Calculer conflits
    targets = defaultdict(list)
    for ged_key in events_ged:
        suggestion, _, _ = suggest_event(ged_key, events_host, events_json, events_tmg)
        if suggestion:
            targets[suggestion].append(ged_key)
    conflicts = {t: keys for t, keys in targets.items() if len(keys) > 1}

    row = 3
    if conflicts:
        for tmg_name, keys in sorted(conflicts.items()):
            norm_target = normalize(tmg_name)
            is_standard_alias = norm_target in keys
            if is_standard_alias:
                fill = FILL_GREEN
                eval_txt = "OK — alias standard (ex: MARR + MARIAGE → même event)"
            else:
                fill = FILL_YELLOW
                eval_txt = "⚠️ Vérifier — plusieurs sources vers même cible"

            ws4.cell(row, 1, tmg_name)
            ws4.cell(row, 2, ", ".join(sorted(keys)))
            ws4.cell(row, 3, eval_txt)
            style_data_row(ws4, row, 3, fill)
            row += 1
    else:
        ws4.cell(3, 1, "Aucun conflit détecté ✓")
        ws4.cell(3, 1).fill = FILL_GREEN
        ws4.cell(3, 1).font = FONT_NORMAL

    ws4.column_dimensions['A'].width = 28
    ws4.column_dimensions['B'].width = 45
    ws4.column_dimensions['C'].width = 58

    # Sauvegarder
    wb.save(EXCEL_FILE)
    log(f"\n✅ Excel generated: {EXCEL_FILE}", 'SUCCESS')
    log(f"   🟢 Green  = validated (previous JSON)", 'INFO')
    log(f"   🟡 Jaune   = suggestion à confirmer", 'INFO')
    log(f"   🔴 Rouge   = à remplir manuellement", 'INFO')
    log(f"   ★ Colonnes 'À VALIDER' = les seules à éditer", 'INFO')


# =============================================================================
# 6. COMPILATION EXCEL → JSON (mode C)
# =============================================================================
def compile_json():
    """Lit l'Excel validé, produit mapping.json."""
    if not os.path.exists(EXCEL_FILE):
        log(f"❌ {EXCEL_FILE} not found. Run in G mode first.", 'ERROR')
        return

    log(f"📖 Reading {EXCEL_FILE}...", 'INFO')
    wb = openpyxl.load_workbook(EXCEL_FILE)

    mapping = {
        '_meta': {
            'generated_at': __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M'),
            'source': EXCEL_FILE,
            'normalize_rule': 'NFKD + ASCII + UPPER + STRIP'
        },
        'events': {},
        'roles':  {}
    }

    errors = []

    # --- EVENTS: colonne G (index 6) = À VALIDER ---
    ws = wb["EVENTS"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        ged_key = row[0].value   # Col A
        valider = row[6].value   # Col G

        if not ged_key or str(ged_key).strip() == "RÉSUMÉ":
            continue

        ged_key = str(ged_key).strip()

        if not valider or str(valider).strip() == "":
            errors.append(f"  🔴 EVENT '{ged_key}': 'TO VALIDATE' column empty")
            continue

        mapping['events'][ged_key] = {'tmg_name': str(valider).strip()}

    # --- ROLES: colonnes H+I (index 7,8) = À VALIDER eng + fra ---
    ws2 = wb["ROLES"]
    for row in ws2.iter_rows(min_row=2, values_only=False):
        ged_key     = row[0].value   # Col A
        valider_eng = row[7].value   # Col H
        valider_fra = row[8].value   # Col I

        if not ged_key or str(ged_key).strip() == "RÉSUMÉ":
            continue

        ged_key = str(ged_key).strip()

        if not valider_eng or str(valider_eng).strip() == "":
            errors.append(f"  🔴 ROLE '{ged_key}': 'TO VALIDATE eng' column empty")
            continue

        mapping['roles'][ged_key] = {
            'eng': str(valider_eng).strip(),
            'fra': str(valider_fra).strip() if valider_fra else str(valider_eng).strip()
        }

    # Écriture JSON
    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False)

    log(f"\n✅ mapping.json écrit: {JSON_FILE}", 'SUCCESS')
    log(f"   {len(mapping['events'])} events, {len(mapping['roles'])} roles", 'INFO')

    if errors:
        log(f"\n⚠️  {len(errors)} entrées incomplètes (non incluses dans le JSON):", 'WARNING')
        for e in errors:
            log(e, 'WARNING')
        log("\n   → Complétez ces lignes dans l'Excel et relancez en mode C.", 'WARNING')
    else:
        log("   ✓ Toutes les entrées complètes — mapping prêt pour l'injecteur.", 'SUCCESS')


# =============================================================================
# FONCTIONS MODE DUAL (CLI / GUI)
# =============================================================================
def generate_excel_mode(gedcom_path=None, tmg_project_path=None, tmg_prefix=None, log_callback=None):
    """
    Génère Excel de mapping - Mode dual CLI/GUI
    
    Mode CLI (standalone):
        python mapping_tool.py G
        → paramètres = None → dialogue GUI de sélection
    
    Mode GUI:
        generate_excel_mode(gedcom, path, prefix, callback)
        → paramètres fournis → pas de dialogue
        → logs vers callback
    
    Returns:
        str: Chemin vers le fichier Excel généré
    """
    global GEDCOM_PATH, TMG_PATH, TMG_PREFIX, LOG_CALLBACK
    
    # Configuration callback
    LOG_CALLBACK = log_callback
    
    # Configuration chemins
    if gedcom_path and tmg_project_path and tmg_prefix:
        # Mode GUI - paramètres fournis
        GEDCOM_PATH = gedcom_path
        TMG_PATH = tmg_project_path
        TMG_PREFIX = tmg_prefix
    else:
        # Mode CLI - dialogue interactif
        select_paths_gui()
    
    # Workflow identique pour CLI et GUI
    log("\n📖 Scan GEDCOM...", 'INFO')
    events_ged, roles_ged = scan_gedcom()
    log(f"   {len(events_ged)} events with _SHAR, {len(roles_ged)} roles", 'INFO')

    log("📖 Scan T.DBF...", 'INFO')
    events_tmg, roles_tmg = scan_tdbf()
    log(f"   {len(events_tmg)} TMG events, {len(roles_tmg)} TMG roles", 'INFO')

    log("📖 Chargement mémoire...", 'INFO')
    events_host = load_host_ddbf()
    events_json, roles_json = load_memory_json()

    log("\n📝 Generating Excel...", 'INFO')
    generate_excel(
        events_ged, roles_ged, events_tmg, roles_tmg,
        events_host, events_json, roles_json
    )
    
    return EXCEL_FILE


def compile_json_mode(excel_file=None, json_file=None, log_callback=None):
    """
    Compile Excel → JSON - Mode dual CLI/GUI
    
    Mode CLI: compile_json_mode()
    Mode GUI: compile_json_mode("mapping_master.xlsx", "mapping.json", callback)
    
    Returns:
        tuple: (json_path, errors_list) ou (None, errors_list) si échec
    """
    global EXCEL_FILE, JSON_FILE, LOG_CALLBACK
    
    # Configuration callback
    LOG_CALLBACK = log_callback
    
    # Configuration fichiers
    if excel_file:
        EXCEL_FILE = excel_file
    if json_file:
        JSON_FILE = json_file
    
    # Vérifier existence Excel
    if not os.path.exists(EXCEL_FILE):
        log(f"❌ {EXCEL_FILE} not found. Run Generate Excel mode first.", 'ERROR')
        return None, [f"File {EXCEL_FILE} not found"]
    
    # Lecture et compilation
    log(f"📖 Reading {EXCEL_FILE}...", 'INFO')
    wb = openpyxl.load_workbook(EXCEL_FILE)

    mapping = {
        '_meta': {
            'generated_at': __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M'),
            'source': EXCEL_FILE,
            'normalize_rule': 'NFKD + ASCII + UPPER + STRIP'
        },
        'events': {},
        'roles':  {}
    }

    errors = []

    # --- EVENTS ---
    ws = wb["EVENTS"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        ged_key = row[0].value
        valider = row[6].value

        if not ged_key or str(ged_key).strip() == "RÉSUMÉ":
            continue

        ged_key = str(ged_key).strip()

        if not valider or str(valider).strip() == "":
            errors.append(f"EVENT '{ged_key}': 'TO VALIDATE' column empty")
            continue

        mapping['events'][ged_key] = {'tmg_name': str(valider).strip()}

    # --- ROLES ---
    ws2 = wb["ROLES"]
    for row in ws2.iter_rows(min_row=2, values_only=False):
        ged_key     = row[0].value
        valider_eng = row[7].value
        valider_fra = row[8].value

        if not ged_key or str(ged_key).strip() == "RÉSUMÉ":
            continue

        ged_key = str(ged_key).strip()

        if not valider_eng or str(valider_eng).strip() == "":
            errors.append(f"ROLE '{ged_key}': 'TO VALIDATE eng' column empty")
            continue

        mapping['roles'][ged_key] = {
            'eng': str(valider_eng).strip(),
            'fra': str(valider_fra).strip() if valider_fra else str(valider_eng).strip()
        }

    # Si erreurs critiques → BLOQUER
    if errors:
        log(f"\n❌ COMPILATION FAILED: {len(errors)} incomplete entries", 'ERROR')
        for e in errors:
            log(f"  🔴 {e}", 'ERROR')
        log("\n   → Complétez ces lignes dans l'Excel et relancez.", 'ERROR')
        return None, errors

    # Écriture JSON
    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, indent=2, ensure_ascii=False)

    log(f"\n✅ mapping.json écrit: {JSON_FILE}", 'SUCCESS')
    log(f"   {len(mapping['events'])} events, {len(mapping['roles'])} roles", 'SUCCESS')
    log("   ✓ Mapping prêt pour l'injecteur", 'SUCCESS')
    
    return JSON_FILE, []


# =============================================================================
# MAIN
# =============================================================================
def main():
    if len(sys.argv) < 2 or sys.argv[1].upper() not in ('G', 'C'):
        print(__doc__)
        print("Usage: python mapping_tool.py G   (génère Excel)")
        print("       python mapping_tool.py C   (compile JSON)")
        sys.exit(0)

    mode = sys.argv[1].upper()

    if mode == 'G':
        log("=" * 70)
        log("MODE G — Excel mapping generation")
        log("=" * 70)
        generate_excel_mode()  # Mode CLI sans paramètres

    elif mode == 'C':
        log("=" * 70)
        log("MODE C — Excel → JSON compilation")
        log("=" * 70)
        compile_json_mode()  # Mode CLI sans paramètres


if __name__ == "__main__":
    main()
